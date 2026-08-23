from typing import List
from openpyxl import load_workbook

from app.core.exceptions import DocumentProcessingException
from .base import DocumentLoader, ParsedSection


class XLSXLoader(DocumentLoader):
    """Loads XLSX files, treating each sheet as a section."""

    def load(self, file_path: str) -> List[ParsedSection]:
        # read_only=True is memory efficient for large files
        wb = load_workbook(filename=file_path, read_only=True, data_only=True)
        sections: List[ParsedSection] = []

        for sheet in wb.sheetnames:
            ws = wb[sheet]
            rows = []
            for row in ws.iter_rows(values_only=True):
                if any(cell is not None for cell in row):
                    # Convert row to CSV-like string
                    row_str = ",".join(str(cell) if cell is not None else "" for cell in row)
                    rows.append(row_str)
            
            if rows:
                sections.append(
                    ParsedSection(
                        content="\n".join(rows),
                        section=sheet
                    )
                )
        
        wb.close()

        if not sections:
            raise DocumentProcessingException("XLSX contained no extractable data.")
            
        return sections